from typing import List, Union

import polars as pl
from openpyxl import load_workbook
from python_calamine import CalamineWorkbook

from utils.logger import logger

from .base_strategy import ReadDataStrategy


def load_column_names(file_path: str, sheet_name: str) -> List[str]:
    workbook = CalamineWorkbook.from_path(file_path)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet_data = sheet.to_python()
    columns = [col for col in sheet_data[0] if col is not None and col]
    return columns


# def load_column_names(file_path: str, sheet_name: str) -> List[str]:
#     wb = load_workbook(file_path, read_only=True)
#     ws = wb[sheet_name]
#     columns = [cell.value for cell in next(ws.iter_rows(max_row=1))]
#     return columns


class ReadExcelFileStrategy(ReadDataStrategy):
    def __init__(self, file_path: str = "") -> None:
        super().__init__(file_path)

    def load(self, sheet_name: Union[str, List[str]], *args, **kwargs):
        usecols = kwargs.get("usecols")
        # Read 0 rows to get column names
        # columns = pd.ExcelFile(self.file_path).parse(sheet_name, nrows=0)
        # columns: List = load_column_names(self.file_path, sheet_name)
        # schema = {col: pl.Utf8 for col in columns if col is not None and col}
        # if usecols is not None:
        #     schema = {col: pl.Utf8 for col in usecols if col in columns}

        df = pl.read_excel(
            self.file_path,
            sheet_name=sheet_name,
            has_header=True,
            # schema_overrides=schema,
            columns=usecols,
            infer_schema_length=0,
            encoding="utf8-lossy",
        )

        if usecols is not None:
            invalid_columns = set(usecols) - set(df.columns)
            if invalid_columns:
                logger.error(
                    f"[{self.__class__.__name__} - {sheet_name}] Column name not in data columns: {', '.join(invalid_columns)}. Data columns: {', '.join(df.columns.value.tolist())}"
                )
                return
            df = df[usecols]

        logger.success(
            f"[{self.__class__.__name__} - {sheet_name}] Data loaded successfully: {df.shape[0]:,} rows, {df.shape[1]:,} columns."
        )
        return df
